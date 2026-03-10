#!/usr/bin/env python3
"""
LG Buying Guides GA Data Parser
Reads monthly xlsx files and master link file, outputs data.json for the dashboard.
"""

import json
import os
import re
import glob
import unicodedata
import openpyxl
from collections import defaultdict

DATA_DIR = os.path.dirname(os.path.abspath(__file__))
MONTH_ORDER = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
MONTH_COL_MAP = {m: i + 8 for i, m in enumerate(MONTH_ORDER)}  # Jan=col8(idx H), Feb=col9(idx I), ...

COUNTRY_MAP = {
    'UK': '영국(UK)', 'FR': '프랑스(FR)', 'PL': '폴란드(PL)', 'ES': '스페인(ES)',
    'BR': '브라질(BR)', 'CA-EN': '캐나다(CA_en)', 'CA-FR': '캐나다(CA_fr)',
    'MX': '멕시코(MX)', 'NL': '네덜란드(NL)', 'IT': '이탈리아(IT)', 'DE': '독일(DE)'
}

# Key metrics and their row definitions (row offsets from year header row=77)
# These are the primary KPIs we extract
KEY_METRICS = {
    # Acquisition
    'page_view_micro': {'section': 'Acquisition', 'label': 'Page View (Micro)', 'row_offset': 3},
    'page_view_lineup': {'section': 'Acquisition', 'label': 'Page View (Lineup)', 'row_offset': 4},
    'session_micro': {'section': 'Acquisition', 'label': 'Session (Micro)', 'row_offset': 6},
    'session_lineup': {'section': 'Acquisition', 'label': 'Session (Lineup)', 'row_offset': 7},
    'session_internal': {'section': 'Acquisition', 'label': 'Internal Sessions', 'row_offset': 8},
    'session_external': {'section': 'Acquisition', 'label': 'External (Entrance)', 'row_offset': 9},
    'session_organic': {'section': 'Acquisition', 'label': 'Organic Sessions', 'row_offset': 10},
    'session_others': {'section': 'Acquisition', 'label': 'Others (Paid/Affiliates/etc)', 'row_offset': 11},
    'engaged_session': {'section': 'Acquisition', 'label': 'Engaged Session', 'row_offset': 12},
    'exit_micro': {'section': 'Acquisition', 'label': 'Exit (Micro)', 'row_offset': 14},
    'exit_lineup': {'section': 'Acquisition', 'label': 'Exit (Lineup)', 'row_offset': 15},
    'pv_per_session_micro': {'section': 'Acquisition', 'label': 'PV/Session (Micro)', 'row_offset': 17, 'is_rate': True},
    'pv_per_session_lineup': {'section': 'Acquisition', 'label': 'PV/Session (Lineup)', 'row_offset': 18, 'is_rate': True},
    'engagement_rate': {'section': 'Acquisition', 'label': 'Engagement Rate', 'row_offset': 19, 'is_rate': True},
    'exit_rate_micro': {'section': 'Acquisition', 'label': 'Exit Rate (Micro)', 'row_offset': 21, 'is_rate': True},
    'exit_rate_lineup': {'section': 'Acquisition', 'label': 'Exit Rate (Lineup)', 'row_offset': 22, 'is_rate': True},

    # Behavior
    'avg_session_duration': {'section': 'Behavior', 'label': 'Avg Session Duration (sec)', 'row_offset': 23},
    'event_click_total': {'section': 'Behavior', 'label': 'Total Event Clicks', 'row_offset': 24},
    'click_gnb_search': {'section': 'Behavior', 'label': 'GNB & Search Clicks', 'row_offset': 25},
    'click_anchor_tab': {'section': 'Behavior', 'label': 'Anchor Tab Clicks', 'row_offset': 26},
    'click_features': {'section': 'Behavior', 'label': 'Features Clicks', 'row_offset': 29},
    'click_per_session_total': {'section': 'Behavior', 'label': 'Total Click/Session', 'row_offset': 99, 'is_rate': True},
    'click_per_session_content': {'section': 'Behavior', 'label': 'Content Click/Session', 'row_offset': 100, 'is_rate': True},

    # Conversion
    'plp_conversion': {'section': 'Conversion', 'label': 'PLP Conversion Rate', 'row_offset': 104, 'is_rate': True},
    'product_conversion': {'section': 'Conversion', 'label': 'Product Conversion Rate', 'row_offset': 105, 'is_rate': True},
    'conv_lineup_sessions': {'section': 'Conversion', 'label': 'Lineup Sessions (Conv)', 'row_offset': 106},
    'conv_plp': {'section': 'Conversion', 'label': 'PLP Visits', 'row_offset': 107},
    'conv_pdp': {'section': 'Conversion', 'label': 'PDP/PBP Visits', 'row_offset': 108},
    'purchase_intent_conversion': {'section': 'Conversion', 'label': 'Purchase Intent Conv Rate', 'row_offset': 109, 'is_rate': True},
    'purchase_conversion': {'section': 'Conversion', 'label': 'Purchase Conv Rate', 'row_offset': 110, 'is_rate': True},
    'conv_add_to_cart': {'section': 'Conversion', 'label': 'Add to Cart / Check-out', 'row_offset': 112},
    'conv_purchase': {'section': 'Conversion', 'label': 'Purchase', 'row_offset': 113},
}


def find_year_row(ws, target_year=2026, max_row=100):
    """Find the row where the year header starts."""
    for row_idx in range(1, max_row + 1):
        cell_val = ws.cell(row=row_idx, column=2).value
        if cell_val == target_year:
            return row_idx
    return None


def extract_metrics_dynamic(ws, year_row):
    """Dynamically extract metrics by scanning the structure from the year row."""
    metrics = {}

    # Build month->column mapping from header row
    month_cols = []
    for col_idx in range(8, 20):  # H through S (Jan-Dec)
        month_name = ws.cell(row=year_row, column=col_idx).value
        if month_name and month_name in MONTH_ORDER:
            month_cols.append((month_name, col_idx))

    # First pass: determine which months have real data by checking key rows
    # (Page View Lineup row, typically 3-4 rows after year header)
    months_with_real_data = set()
    for check_offset in [3, 4, 5, 6, 7]:  # check several rows
        check_row = year_row + check_offset
        for month_name, col_idx in month_cols:
            val = ws.cell(row=check_row, column=col_idx).value
            if val is not None and val != 'N/A' and val != 0 and val != '#DIV/0!':
                try:
                    fval = float(val)
                    if fval > 0:
                        months_with_real_data.add(month_name)
                except (ValueError, TypeError):
                    pass

    # Only use months that have real data
    active_month_cols = [(m, c) for m, c in month_cols if m in months_with_real_data]

    # Scan rows after year header to build metric structure
    current_section = None
    row_idx = year_row + 1
    months_with_data = set()

    while row_idx <= ws.max_row:
        col_b = ws.cell(row=row_idx, column=2).value
        col_c = ws.cell(row=row_idx, column=3).value

        # Check if we hit next year section or empty gap
        if col_b and isinstance(col_b, (int, float)):
            break

        # Section headers: Acquisition, Behavior, Conversion
        if col_b in ('Acquisition', 'Behavior', 'Conversion'):
            current_section = col_b

        # Extract metric name from column C (or B for section-level)
        metric_name = None
        if col_c and isinstance(col_c, str):
            metric_name = col_c
        elif col_b and isinstance(col_b, str) and col_b not in ('Acquisition', 'Behavior', 'Conversion'):
            metric_name = col_b

        if metric_name and current_section:
            raw_name = metric_name
            monthly_data = {}
            for month_name, col_idx in active_month_cols:
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None and val != 'N/A':
                    try:
                        monthly_data[month_name] = float(val)
                        months_with_data.add(month_name)
                    except (ValueError, TypeError):
                        pass

            if monthly_data:
                # Create a clean key
                stripped = raw_name.strip()
                clean_key = re.sub(r'[^\w\s]', '', stripped).strip().lower().replace(' ', '_')
                indent_level = len(raw_name) - len(raw_name.lstrip())

                # Handle duplicate keys by appending section
                if clean_key in metrics:
                    clean_key = f"{clean_key}_{current_section.lower()}"

                metrics[clean_key] = {
                    'section': current_section,
                    'label': stripped,
                    'indent': indent_level // 5,
                    'monthly': monthly_data
                }

        row_idx += 1

    # Only return months that actually have data
    valid_months = [m for m, _ in month_cols if m in months_with_data]
    return metrics, valid_months


def parse_master_file():
    """Parse the master link file for page status info."""
    master_files = glob.glob(os.path.join(DATA_DIR, 'LG_Buying_Guides_Master*Link*.xlsx'))
    if not master_files:
        return []

    wb = openpyxl.load_workbook(master_files[0], data_only=True)
    ws = wb[wb.sheetnames[0]]

    pages = []
    current_country = None

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        vals = [cell.value for cell in row]

        # Country in column B (index 1)
        if vals[1]:
            current_country = str(vals[1]).strip()

        category = vals[2]  # 구분 (TV, 모니터, etc.)
        detail = vals[3]    # 상세 (Lineup Guide, Feature Library)
        link = vals[4]      # 링크
        feedback = vals[5]  # 피드백 반영 여부
        ga_tag = vals[6]    # GA 태깅
        live_status = vals[7]  # 라이브 상태
        breadcrumb = vals[8]   # 브레드크럼 노출

        if detail and link:
            pages.append({
                'country': current_country or '',
                'category': str(category).strip() if category else '',
                'type': str(detail).strip(),
                'link': str(link).strip(),
                'feedback': str(feedback).strip() if feedback else '',
                'ga_tagged': str(ga_tag).strip() if ga_tag else '',
                'live_status': str(live_status).strip() if live_status else '',
                'breadcrumb': str(breadcrumb).strip() if breadcrumb else '',
            })

    return pages


CONTENT_TYPES = {
    '라인업가이드': 'lineup',
    '스펙라이브러리': 'feature_library',
}


def find_monthly_files_by_type():
    """Find monthly files grouped by content type and sorted by month."""
    from pathlib import Path
    files = [str(f) for f in Path(DATA_DIR).glob('Monthly*.xlsx')]

    result = {}  # content_type -> list of (filepath, month_idx, month_name)
    for f in files:
        basename = unicodedata.normalize('NFC', os.path.basename(f))
        # Determine content type
        ctype = 'lineup'  # default
        for kr_name, eng_type in CONTENT_TYPES.items():
            if kr_name in basename:
                ctype = eng_type
                break

        for i, month in enumerate(MONTH_ORDER):
            if month in basename:
                result.setdefault(ctype, []).append((f, i, month))
                break

    # Sort each type by month
    for ctype in result:
        result[ctype].sort(key=lambda x: x[1])

    return result


def parse_single_workbook(filepath, content_type):
    """Parse a single workbook (lineup or feature library)."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    all_data = {}
    all_months = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        parts = sheet_name.split('_', 1)
        if len(parts) != 2:
            continue

        category = parts[0]  # TV or Monitor
        country_code = parts[1]  # UK, FR, etc.

        year_row = find_year_row(ws, 2026)
        if not year_row:
            print(f"  Warning: No 2026 data found in {sheet_name}")
            continue

        metrics, months_2026 = extract_metrics_dynamic(ws, year_row)
        all_months.update(months_2026)

        year_row_2025 = find_year_row(ws, 2025)
        metrics_2025 = {}
        if year_row_2025:
            metrics_2025, _ = extract_metrics_dynamic(ws, year_row_2025)

        # Extract Data Insights text
        insights_text = []
        for row_idx in range(18, 30):
            cell_val = ws.cell(row=row_idx, column=2).value
            if cell_val and isinstance(cell_val, str):
                text = cell_val.strip()
                if text.startswith('<Data Insights>'):
                    continue
                col_c = ws.cell(row=row_idx, column=3).value
                if col_c:
                    text = text + ' ' + str(col_c).strip()
                if text:
                    insights_text.append(text)

        # Use content_type prefix to avoid key collisions
        data_key = f"{content_type}_{sheet_name}"
        all_data[data_key] = {
            'category': category,
            'country': country_code,
            'content_type': content_type,
            'metrics_2026': metrics,
            'metrics_2025': metrics_2025,
            'insights': insights_text,
        }

    return all_data, all_months


def parse_monthly_data():
    """Parse all monthly files (lineup + feature library)."""
    files_by_type = find_monthly_files_by_type()
    if not files_by_type:
        return {}, []

    all_data = {}
    all_months = set()

    for ctype, file_list in files_by_type.items():
        latest_file, _, latest_month = file_list[-1]
        print(f"  Reading {ctype}: {os.path.basename(latest_file)} (up to {latest_month})")

        data, months = parse_single_workbook(latest_file, ctype)
        all_data.update(data)
        all_months.update(months)

    sorted_months = sorted(all_months, key=lambda m: MONTH_ORDER.index(m))
    return all_data, sorted_months


def _get_metric(metrics, *keywords):
    """Find metric value helper."""
    for k, v in metrics.items():
        lk = k.lower()
        if all(kw in lk for kw in keywords):
            return v
    return None


def _get_val(metrics, month, *keywords):
    m = _get_metric(metrics, *keywords)
    if m and month in m.get('monthly', {}):
        return m['monthly'][month]
    return None


def _get_session_val(metrics, month):
    """Get session value - works for both lineup and spec library."""
    return (_get_val(metrics, month, 'lineup', 'session')
            or _get_val(metrics, month, 'tv_lineup', 'session')
            or _get_val(metrics, month, 'monitor_lineup', 'session')
            or _get_val(metrics, month, 'spec_library', 'session')
            or _get_val(metrics, month, 'tv_spec', 'session')
            or _get_val(metrics, month, 'monitor_spec', 'session'))


def _safe_pct(cur, prev):
    if prev and prev > 0 and cur is not None:
        return ((cur - prev) / prev) * 100
    return None


def compute_insights(all_data, months):
    """Generate expert-level strategic insights from all data points."""
    insights = []
    if len(months) < 2:
        return insights

    latest, prev = months[-1], months[-2]

    # ── Collect sheet-level summaries ──
    summaries = []
    for sheet_key, data in all_data.items():
        m = data['metrics_2026']
        ct = data.get('content_type', 'lineup')
        ctype_label = 'Feature Library' if ct == 'feature_library' else 'Lineup Guide'
        label = f"{data['category']} {data['country']} ({ctype_label})"
        cat, country = data['category'], data['country']

        s = {
            'key': sheet_key, 'label': label, 'cat': cat, 'country': country, 'ct': ct,
            'sess': _get_session_val(m, latest),
            'sess_prev': _get_session_val(m, prev),
            'organic': _get_val(m, latest, 'organic'),
            'organic_prev': _get_val(m, prev, 'organic'),
            'internal': _get_val(m, latest, 'internal'),
            'internal_prev': _get_val(m, prev, 'internal'),
            'external': _get_val(m, latest, 'external'),
            'external_prev': _get_val(m, prev, 'external'),
            'clicks': _get_val(m, latest, 'event_click') or _get_val(m, latest, 'guide_event_click') or _get_val(m, latest, 'library_event_click'),
            'clicks_prev': _get_val(m, prev, 'event_click') or _get_val(m, prev, 'guide_event_click') or _get_val(m, prev, 'library_event_click'),
            'duration': _get_val(m, latest, 'avg_session_duration') or _get_val(m, latest, 'session_duration'),
            'duration_prev': _get_val(m, prev, 'avg_session_duration') or _get_val(m, prev, 'session_duration'),
            'plp': _get_val(m, latest, 'plp_conversion'),
            'plp_prev': _get_val(m, prev, 'plp_conversion'),
            'purchase_conv': _get_val(m, latest, 'purchase_conversion'),
            'purchase_conv_prev': _get_val(m, prev, 'purchase_conversion'),
            'engagement': _get_val(m, latest, 'engagem'),
            'engagement_prev': _get_val(m, prev, 'engagem'),
            'exit_rate': _get_val(m, latest, 'exit_rate'),
        }
        summaries.append(s)

    # ── Per-sheet metric insights ──
    for s in summaries:
        label = s['label']
        ct = s['ct']

        # Session changes
        sess_ch = _safe_pct(s['sess'], s['sess_prev'])
        if sess_ch is not None and abs(sess_ch) > 15:
            if sess_ch < -30:
                # Diagnose WHY
                int_ch = _safe_pct(s['internal'], s['internal_prev'])
                org_ch = _safe_pct(s['organic'], s['organic_prev'])
                cause = []
                if int_ch is not None and int_ch < -20:
                    cause.append(f"internal traffic down {int_ch:.0f}%")
                if org_ch is not None and org_ch < -20:
                    cause.append(f"organic down {org_ch:.0f}%")
                cause_str = f" Driven by: {', '.join(cause)}." if cause else ""
                insights.append({'type': 'warning', 'page': label, 'metric': 'Sessions', 'content_type': ct, 'priority': 'high',
                    'message': f"{label}: Sessions dropped {abs(sess_ch):.0f}% ({s['sess_prev']:.0f}→{s['sess']:.0f}).{cause_str} Investigate internal link exposure, seasonal pattern, or page discoverability."})
            elif sess_ch > 20:
                insights.append({'type': 'success', 'page': label, 'metric': 'Sessions', 'content_type': ct, 'priority': 'medium',
                    'message': f"{label}: Sessions grew {sess_ch:.0f}% ({s['sess_prev']:.0f}→{s['sess']:.0f}). Positive momentum — review what's working and amplify."})

        # PLP Conversion
        plp_ch = _safe_pct(s['plp'], s['plp_prev'])
        if plp_ch is not None and s['plp'] is not None and s['plp'] <= 1:
            if plp_ch < -20 and s['plp_prev'] and s['plp_prev'] > 0.3:
                insights.append({'type': 'warning', 'page': label, 'metric': 'PLP Conversion', 'content_type': ct, 'priority': 'high',
                    'message': f"{label}: PLP conversion fell from {s['plp_prev']:.1%} to {s['plp']:.1%} ({plp_ch:.0f}%). CTA effectiveness declining — review product link placement and above-fold messaging."})
            elif s['plp'] and s['plp'] > 0.5 and (s['sess'] or 0) > 20:
                insights.append({'type': 'success', 'page': label, 'metric': 'PLP Conversion', 'content_type': ct, 'priority': 'medium',
                    'message': f"{label}: Strong PLP conversion at {s['plp']:.1%} with {s['sess']:.0f} sessions. High-performing content — benchmark for other markets."})

        # Content engagement quality
        if s['clicks'] and s['sess'] and s['sess'] > 10:
            clicks_per_sess = s['clicks'] / s['sess']
            if clicks_per_sess < 1.5 and s['duration'] and s['duration'] > 120:
                insights.append({'type': 'info', 'page': label, 'metric': 'Engagement', 'content_type': ct, 'priority': 'medium',
                    'message': f"{label}: High dwell time ({s['duration']:.0f}s) but low click rate ({clicks_per_sess:.1f}/session). Users are reading but not interacting — content may lack clear CTAs or interactive elements."})
            elif clicks_per_sess > 5:
                insights.append({'type': 'success', 'page': label, 'metric': 'Engagement', 'content_type': ct, 'priority': 'low',
                    'message': f"{label}: Excellent interaction rate ({clicks_per_sess:.1f} clicks/session). Users actively exploring features and products."})

        # Organic dependency / opportunity
        if s['external'] and s['external'] > 0 and s['organic'] is not None:
            org_share = s['organic'] / s['external'] if s['external'] > 0 else 0
            if org_share < 0.2 and (s['sess'] or 0) > 20:
                insights.append({'type': 'warning', 'page': label, 'metric': 'SEO', 'content_type': ct, 'priority': 'medium',
                    'message': f"{label}: Organic makes up only {org_share:.0%} of external traffic. Heavy reliance on paid/direct — SEO investment needed for sustainable growth."})

        # Purchase funnel
        if s['plp'] and s['purchase_conv'] is not None and s['sess'] and s['sess'] > 20:
            if s['plp'] > 0.3 and s['purchase_conv'] < 0.005:
                insights.append({'type': 'info', 'page': label, 'metric': 'Funnel', 'content_type': ct, 'priority': 'medium',
                    'message': f"{label}: Good PLP conversion ({s['plp']:.1%}) but purchase conversion is near zero ({s['purchase_conv']:.2%}). Content drives product discovery but not purchase intent — consider adding pricing context, promotions, or urgency cues."})

    # ── Cross-cutting strategic insights ──
    # Lineup vs Feature Library comparison
    lineup_sheets = [s for s in summaries if s['ct'] == 'lineup']
    fl_sheets = [s for s in summaries if s['ct'] == 'feature_library']

    if lineup_sheets and fl_sheets:
        lu_avg_plp = [s['plp'] for s in lineup_sheets if s['plp'] and s['plp'] <= 1]
        fl_avg_plp = [s['plp'] for s in fl_sheets if s['plp'] and s['plp'] <= 1]
        if lu_avg_plp and fl_avg_plp:
            lu_mean = sum(lu_avg_plp) / len(lu_avg_plp)
            fl_mean = sum(fl_avg_plp) / len(fl_avg_plp)
            if lu_mean > fl_mean * 1.3:
                insights.append({'type': 'info', 'page': 'Cross-Content', 'metric': 'Content Strategy', 'content_type': 'all', 'priority': 'high',
                    'message': f"Lineup Guides convert {lu_mean:.1%} to PLP vs Feature Library's {fl_mean:.1%}. Lineup content is {((lu_mean/fl_mean - 1)*100):.0f}% more effective at driving product page visits. Feature Library needs stronger product CTAs or inline product links to close this gap."})

        lu_avg_dur = [s['duration'] for s in lineup_sheets if s['duration'] and s['duration'] > 0]
        fl_avg_dur = [s['duration'] for s in fl_sheets if s['duration'] and s['duration'] > 0]
        if lu_avg_dur and fl_avg_dur:
            lu_dur = sum(lu_avg_dur) / len(lu_avg_dur)
            fl_dur = sum(fl_avg_dur) / len(fl_avg_dur)
            insights.append({'type': 'info', 'page': 'Cross-Content', 'metric': 'Content Strategy', 'content_type': 'all', 'priority': 'medium',
                'message': f"Avg dwell time: Lineup {lu_dur:.0f}s vs Feature Library {fl_dur:.0f}s. {'Feature Library commands more attention per visit — leverage this engagement depth with richer product integration.' if fl_dur > lu_dur else 'Lineup Guide holds attention longer — this is the primary content discovery touchpoint.'}"})

    # Top market patterns
    top_markets = {}  # country -> aggregated
    for s in summaries:
        c = s['country']
        if c not in top_markets:
            top_markets[c] = {'total_sess': 0, 'total_sess_prev': 0, 'plps': [], 'purchases': 0, 'ctypes': set()}
        top_markets[c]['total_sess'] += s['sess'] or 0
        top_markets[c]['total_sess_prev'] += s['sess_prev'] or 0
        if s['plp'] and s['plp'] <= 1:
            top_markets[c]['plps'].append(s['plp'])
        if s['purchase_conv']:
            top_markets[c]['purchases'] += 1
        top_markets[c]['ctypes'].add(s['ct'])

    # Market concentration risk
    total_all = sum(v['total_sess'] for v in top_markets.values())
    if total_all > 0:
        sorted_markets = sorted(top_markets.items(), key=lambda x: x[1]['total_sess'], reverse=True)
        top_share = sorted_markets[0][1]['total_sess'] / total_all if total_all > 0 else 0
        if top_share > 0.3:
            insights.insert(0, {'type': 'info', 'page': 'Global Strategy', 'metric': 'Market Strategy', 'content_type': 'all', 'priority': 'high',
                'message': f"{sorted_markets[0][0]} accounts for {top_share:.0%} of all sessions. Over-concentration risk — invest in growing underperforming markets ({', '.join(c for c,v in sorted_markets[-3:] if v['total_sess'] > 0)}) to diversify traffic sources."})

    # Markets with best conversion efficiency
    efficient = [(c, sum(v['plps'])/len(v['plps']), v['total_sess'])
                 for c, v in top_markets.items() if len(v['plps']) >= 2 and v['total_sess'] > 30]
    efficient.sort(key=lambda x: x[1], reverse=True)
    if efficient:
        best = efficient[0]
        insights.insert(0, {'type': 'success', 'page': 'Global Strategy', 'metric': 'Conversion Efficiency', 'content_type': 'all', 'priority': 'high',
            'message': f"{best[0]} leads in conversion efficiency ({best[1]:.1%} avg PLP conv, {best[2]:.0f} sessions). Study this market's content structure and user journey as a template for underperformers."})

    return insights


def compute_strategic_narrative(all_data, months, pages):
    """Generate high-level strategic narratives for the content team."""
    if len(months) < 2:
        return {}

    latest, prev = months[-1], months[-2]
    narratives = {}

    # Collect all summaries
    summaries = []
    for sheet_key, data in all_data.items():
        m = data['metrics_2026']
        ct = data.get('content_type', 'lineup')
        s = {
            'ct': ct, 'cat': data['category'], 'country': data['country'],
            'sess': _get_session_val(m, latest),
            'sess_prev': _get_session_val(m, prev),
            'plp': _get_val(m, latest, 'plp_conversion'),
            'clicks': _get_val(m, latest, 'event_click') or _get_val(m, latest, 'guide_event_click') or _get_val(m, latest, 'library_event_click'),
            'duration': _get_val(m, latest, 'avg_session_duration') or _get_val(m, latest, 'session_duration'),
            'organic': _get_val(m, latest, 'organic'),
            'external': _get_val(m, latest, 'external'),
            'internal': _get_val(m, latest, 'internal'),
            'purchase_conv': _get_val(m, latest, 'purchase_conversion'),
            'engagement': _get_val(m, latest, 'engagem'),
        }
        summaries.append(s)

    lu = [s for s in summaries if s['ct'] == 'lineup']
    fl = [s for s in summaries if s['ct'] == 'feature_library']

    # ── Content Team Direction ──
    content_actions = []

    # 1. Lineup vs Feature Library strategy
    lu_total = sum(s['sess'] or 0 for s in lu)
    fl_total = sum(s['sess'] or 0 for s in fl)
    lu_plp = [s['plp'] for s in lu if s['plp'] and s['plp'] <= 1]
    fl_plp = [s['plp'] for s in fl if s['plp'] and s['plp'] <= 1]
    lu_avg_plp = sum(lu_plp)/len(lu_plp) if lu_plp else 0
    fl_avg_plp = sum(fl_plp)/len(fl_plp) if fl_plp else 0

    content_actions.append({
        'title': 'Lineup Guide vs Feature Library: Role Clarity',
        'title_ko': '라인업 가이드 vs 피쳐 라이브러리: 역할 정의',
        'body': f"Lineup Guides drive {lu_total:.0f} sessions ({lu_avg_plp:.1%} PLP conv) vs Feature Library's {fl_total:.0f} sessions ({fl_avg_plp:.1%} PLP conv). Lineup is the primary purchase-path content; Feature Library serves as a spec-reference tool. Action: Strengthen cross-linking between them — add 'Compare detailed specs' CTAs in Lineup pages, and 'Find your ideal model' CTAs in Feature Library.",
        'body_ko': f"라인업 가이드: {lu_total:.0f} 세션 (PLP 전환 {lu_avg_plp:.1%}), 피쳐 라이브러리: {fl_total:.0f} 세션 (PLP 전환 {fl_avg_plp:.1%}). 라인업은 구매 경로의 핵심 콘텐츠이고, 피쳐 라이브러리는 스펙 참조 도구 역할. 액션: 라인업 페이지에 '상세 스펙 비교하기' CTA를, 피쳐 라이브러리에 '나에게 맞는 모델 찾기' CTA를 추가하여 상호 연결 강화.",
        'priority': 'high',
    })

    # 2. Market-tier strategy
    by_country = {}
    for s in summaries:
        c = s['country']
        if c not in by_country:
            by_country[c] = {'sess': 0, 'plps': [], 'has_purchase': False}
        by_country[c]['sess'] += s['sess'] or 0
        if s['plp'] and s['plp'] <= 1:
            by_country[c]['plps'].append(s['plp'])
        if s['purchase_conv'] and s['purchase_conv'] > 0.005:
            by_country[c]['has_purchase'] = True

    tier1 = [(c, d) for c, d in by_country.items() if d['sess'] > 100]
    tier2 = [(c, d) for c, d in by_country.items() if 30 < d['sess'] <= 100]
    tier3 = [(c, d) for c, d in by_country.items() if d['sess'] <= 30 and d['sess'] > 0]

    t1_names = ', '.join(c for c,_ in sorted(tier1, key=lambda x: x[1]['sess'], reverse=True))
    t2_names = ', '.join(c for c,_ in sorted(tier2, key=lambda x: x[1]['sess'], reverse=True))
    t3_names = ', '.join(c for c,_ in sorted(tier3, key=lambda x: x[1]['sess'], reverse=True))

    content_actions.append({
        'title': 'Three-Tier Market Approach',
        'title_ko': '3단계 시장 전략',
        'body': f"Tier 1 (Scale): {t1_names or 'None'} — Focus on conversion optimization, A/B testing, SEO authority building. Tier 2 (Grow): {t2_names or 'None'} — Increase internal linking exposure, localize content depth, improve discoverability. Tier 3 (Seed): {t3_names or 'None'} — Foundational SEO, ensure basic content quality, monitor for traction signals before heavy investment.",
        'body_ko': f"Tier 1 (확대): {t1_names or '없음'} — 전환율 최적화, A/B 테스트, SEO 권위 구축 집중. Tier 2 (성장): {t2_names or '없음'} — 내부 링킹 노출 확대, 콘텐츠 현지화 심화, 발견성 개선. Tier 3 (씨앗): {t3_names or '없음'} — 기본 SEO 작업, 콘텐츠 품질 보장, 의미있는 트래픽 신호 감지 후 투자 확대.",
        'priority': 'high',
    })

    # 3. Organic growth opportunity
    low_organic = []
    for s in summaries:
        if s['external'] and s['external'] > 5 and s['organic'] is not None:
            org_share = s['organic'] / s['external'] if s['external'] > 0 else 0
            if org_share < 0.3:
                low_organic.append(f"{s['cat']} {s['country']} ({s['ct']})")

    if low_organic:
        content_actions.append({
            'title': 'SEO/Organic Growth Strategy',
            'title_ko': 'SEO/오가닉 성장 전략',
            'body': f"Pages with low organic share (<30%): {', '.join(low_organic[:5])}. These pages rely heavily on internal/paid traffic. Actions: (1) Optimize H1/meta titles with buying-intent keywords (e.g., 'best LG TV 2026', 'LG OLED vs QNED comparison'), (2) Add structured data (FAQ, Product schema), (3) Build internal linking from high-traffic category pages.",
            'body_ko': f"오가닉 비율 낮은 페이지 (<30%): {', '.join(low_organic[:5])}. 내부/유료 트래픽에 과도하게 의존. 액션: (1) 구매 의도 키워드로 H1/메타 타이틀 최적화 (예: 'best LG TV 2026', 'LG OLED vs QNED 비교'), (2) 구조화 데이터 추가 (FAQ, Product 스키마), (3) 고트래픽 카테고리 페이지에서 내부 링킹 강화.",
            'priority': 'high',
        })

    # 4. Content engagement optimization
    low_click_high_dur = []
    for s in summaries:
        if s['sess'] and s['sess'] > 15 and s['clicks'] and s['duration']:
            cps = s['clicks'] / s['sess']
            if cps < 2.0 and s['duration'] > 100:
                low_click_high_dur.append(f"{s['cat']} {s['country']} ({s['ct']}, {cps:.1f} clicks/sess, {s['duration']:.0f}s)")

    if low_click_high_dur:
        content_actions.append({
            'title': 'Engagement Gap: Reading Without Acting',
            'title_ko': '참여 격차: 읽기는 하지만 행동하지 않음',
            'body': f"These pages have high dwell time but low interaction: {', '.join(low_click_high_dur[:4])}. Users consume content but don't click through. Actions: (1) Add in-content product cards with direct PLP links, (2) Insert comparison tables with 'See price' buttons, (3) Use sticky footer CTA bar, (4) Add anchor navigation for faster spec browsing.",
            'body_ko': f"높은 체류시간 대비 낮은 상호작용: {', '.join(low_click_high_dur[:4])}. 사용자가 콘텐츠를 소비하지만 클릭하지 않음. 액션: (1) PLP 직접 링크가 있는 인콘텐츠 제품 카드 추가, (2) '가격 보기' 버튼이 있는 비교 테이블 삽입, (3) 스티키 하단 CTA 바 사용, (4) 빠른 스펙 탐색을 위한 앵커 네비게이션 추가.",
            'priority': 'medium',
        })

    # 5. Purchase funnel gap
    high_plp_low_purchase = []
    for s in summaries:
        if s['plp'] and s['plp'] > 0.3 and s['sess'] and s['sess'] > 20:
            if not s['purchase_conv'] or s['purchase_conv'] < 0.005:
                high_plp_low_purchase.append(f"{s['cat']} {s['country']} ({s['ct']}, PLP {s['plp']:.0%})")

    if high_plp_low_purchase:
        content_actions.append({
            'title': 'Purchase Funnel Leakage',
            'title_ko': '구매 퍼널 누수',
            'body': f"Good product discovery but near-zero purchase: {', '.join(high_plp_low_purchase[:4])}. The content successfully drives users to product pages, but they drop off before purchase. This may be a PDP/checkout issue rather than a content issue. Actions: (1) Coordinate with e-commerce team on PDP conversion, (2) Add promo/bundle offers within buying guide content, (3) Test 'Add to Cart' CTAs directly in lineup comparisons.",
            'body_ko': f"제품 발견은 잘되지만 구매는 거의 없음: {', '.join(high_plp_low_purchase[:4])}. 콘텐츠가 상품 페이지로의 이동은 성공적이나 구매 전 이탈. 콘텐츠보다 PDP/체크아웃 문제일 수 있음. 액션: (1) 이커머스팀과 PDP 전환율 개선 협의, (2) 바잉가이드 내 프로모/번들 오퍼 추가, (3) 라인업 비교 내 '장바구니 담기' CTA 직접 테스트.",
            'priority': 'medium',
        })

    # 6. TV vs Monitor strategy
    tv_sess = sum(s['sess'] or 0 for s in summaries if s['cat'] == 'TV')
    mon_sess = sum(s['sess'] or 0 for s in summaries if s['cat'] == 'Monitor')
    tv_plp = [s['plp'] for s in summaries if s['cat'] == 'TV' and s['plp'] and s['plp'] <= 1]
    mon_plp = [s['plp'] for s in summaries if s['cat'] == 'Monitor' and s['plp'] and s['plp'] <= 1]
    tv_avg = sum(tv_plp)/len(tv_plp) if tv_plp else 0
    mon_avg = sum(mon_plp)/len(mon_plp) if mon_plp else 0

    content_actions.append({
        'title': 'TV vs Monitor: Category-Specific Strategy',
        'title_ko': 'TV vs Monitor: 카테고리별 전략',
        'body': f"TV: {tv_sess:.0f} sessions, {tv_avg:.1%} avg PLP conv. Monitor: {mon_sess:.0f} sessions, {mon_avg:.1%} avg PLP conv. {'TV drives significantly more traffic and should remain the priority content investment. Monitor content needs stronger internal link placement from product category pages to grow its audience.' if tv_sess > mon_sess * 2 else 'Both categories show comparable engagement. Ensure equal content quality and feature freshness across both.'}",
        'body_ko': f"TV: {tv_sess:.0f} 세션, PLP 전환 {tv_avg:.1%}. Monitor: {mon_sess:.0f} 세션, PLP 전환 {mon_avg:.1%}. {'TV가 훨씬 더 많은 트래픽을 유도하며 우선 콘텐츠 투자가 필요. Monitor 콘텐츠는 카테고리 페이지에서의 내부 링크 배치 강화로 오디언스 확대 필요.' if tv_sess > mon_sess * 2 else '두 카테고리 모두 비슷한 참여도를 보임. 양쪽 모두 콘텐츠 품질과 피쳐 최신성을 동일하게 유지.'}",
        'priority': 'medium',
    })

    narratives['content_actions'] = content_actions
    narratives['latest_month'] = latest
    narratives['prev_month'] = prev
    narratives['total_sessions'] = sum(s['sess'] or 0 for s in summaries)
    narratives['total_pages'] = len(summaries)

    return narratives


def compute_expert_analysis(all_data, months, pages):
    """Generate comprehensive expert-level analysis for both lineup and feature library."""
    if not months:
        return {}

    latest = months[-1]
    prev = months[-2] if len(months) > 1 else None

    # ── Build per-sheet summary ──
    sheet_summaries = {}
    for sheet_key, data in all_data.items():
        m = data['metrics_2026']
        ctype = data.get('content_type', 'lineup')

        s = {
            'content_type': ctype,
            'sessions': _get_session_val(m, latest),
            'sessions_prev': _get_session_val(m, prev) if prev else None,
            'page_views': _get_val(m, latest, 'line_up') or _get_val(m, latest, 'lineup') or _get_val(m, latest, 'spec_library'),
            'engaged': _get_val(m, latest, 'engaged'),
            'event_clicks': _get_val(m, latest, 'event_click') or _get_val(m, latest, 'guide_event_click') or _get_val(m, latest, 'library_event_click'),
            'event_clicks_prev': (_get_val(m, prev, 'event_click') or _get_val(m, prev, 'guide_event_click') or _get_val(m, prev, 'library_event_click')) if prev else None,
            'duration': _get_val(m, latest, 'avg_session_duration') or _get_val(m, latest, 'session_duration'),
            'duration_prev': (_get_val(m, prev, 'avg_session_duration') or _get_val(m, prev, 'session_duration')) if prev else None,
            'plp_conv': _get_val(m, latest, 'plp_conversion'),
            'plp_conv_prev': _get_val(m, prev, 'plp_conversion') if prev else None,
            'product_conv': _get_val(m, latest, 'product_conversion'),
            'purchase_conv': _get_val(m, latest, 'purchase_conversion'),
            'purchase_conv_prev': _get_val(m, prev, 'purchase_conversion') if prev else None,
            'purchase_count': _get_val(m, latest, 'purchase') if _get_val(m, latest, 'purchase') and _get_val(m, latest, 'purchase') < 1000 else None,
            'organic': _get_val(m, latest, 'organic'),
            'external': _get_val(m, latest, 'external'),
            'internal': _get_val(m, latest, 'internal'),
            'engagement_rate': _get_val(m, latest, 'engagem'),
            'exit_rate': _get_val(m, latest, 'exit_rate'),
        }
        if s['page_views'] and s['page_views'] > 5000:
            s['page_views'] = None
        sheet_summaries[sheet_key] = s

    # ── Build reports per content type ──
    content_types_found = sorted(set(d.get('content_type', 'lineup') for d in all_data.values()))

    def _build_reports_for(ct_filter=None):
        """Build country/category/executive for a content type (or all)."""
        filtered_data = {k: v for k, v in all_data.items()
                        if ct_filter is None or v.get('content_type', 'lineup') == ct_filter}
        filtered_sums = {k: v for k, v in sheet_summaries.items() if k in filtered_data}

        # Country reports
        cr = {}
        countries = sorted(set(d['country'] for d in filtered_data.values()))
        for country in countries:
            sheets_c = {k: v for k, v in filtered_data.items() if v['country'] == country}
            sums_c = {k: filtered_sums[k] for k in sheets_c if k in filtered_sums}

            total_sessions = sum(s.get('sessions') or 0 for s in sums_c.values())
            total_sessions_prev = sum(s.get('sessions_prev') or 0 for s in sums_c.values()) if prev else 0
            total_clicks = sum(s.get('event_clicks') or 0 for s in sums_c.values())
            avg_plp = [s['plp_conv'] for s in sums_c.values() if s.get('plp_conv') is not None]
            avg_plp_val = sum(avg_plp)/len(avg_plp) if avg_plp else None
            avg_purch = [s['purchase_conv'] for s in sums_c.values() if s.get('purchase_conv') is not None]
            avg_purch_val = sum(avg_purch)/len(avg_purch) if avg_purch else None
            total_purchases = sum(s.get('purchase_count') or 0 for s in sums_c.values())
            avg_dur = [s['duration'] for s in sums_c.values() if s.get('duration')]
            avg_dur_val = sum(avg_dur)/len(avg_dur) if avg_dur else None

            sess_change = ((total_sessions - total_sessions_prev) / total_sessions_prev * 100) if total_sessions_prev > 0 else None
            status = 'stable'
            if sess_change is not None:
                if sess_change > 10: status = 'growing'
                elif sess_change < -20: status = 'declining'

            country_pages = [p for p in pages if country in p.get('country', '')]
            country_kr = COUNTRY_MAP.get(country, country)
            if not country_pages:
                country_pages = [p for p in pages if country_kr and any(c in p.get('country', '') for c in [country_kr, country])]

            ctype_label = {'lineup': 'Lineup Guide', 'feature_library': 'Spec Library'}.get(ct_filter, '')
            analyst_notes = []
            for k, sd in sheets_c.items():
                if sd.get('insights'):
                    meaningful = [n for n in sd['insights'] if len(str(n)) > 20]
                    if meaningful:
                        label = f"{sd['category']} {ctype_label}" if ctype_label else f"{sd['category']}"
                        analyst_notes.append({'page': label, 'notes': meaningful})

            per_product = {}
            for k, sd in sheets_c.items():
                s = sums_c.get(k, {})
                per_product[sd['category']] = {
                    'sessions': s.get('sessions'),
                    'page_views': s.get('page_views'),
                    'plp_conv': s.get('plp_conv'),
                    'purchase_conv': s.get('purchase_conv'),
                    'duration': s.get('duration'),
                    'clicks': s.get('event_clicks'),
                    'engagement_rate': s.get('engagement_rate'),
                }

            cr[country] = {
                'total_sessions': total_sessions,
                'session_change_pct': round(sess_change, 1) if sess_change else None,
                'total_clicks': total_clicks,
                'avg_plp_conv': round(avg_plp_val, 4) if avg_plp_val else None,
                'avg_purchase_conv': round(avg_purch_val, 4) if avg_purch_val else None,
                'total_purchases': total_purchases,
                'avg_duration': round(avg_dur_val, 1) if avg_dur_val else None,
                'status': status,
                'links': country_pages,
                'analyst_notes': analyst_notes,
                'per_product': per_product,
            }

        # Category reports
        cat_reps = {}
        categories = sorted(set(d['category'] for d in filtered_data.values()))
        for cat in categories:
            cat_sheets = {k: v for k, v in filtered_data.items() if v['category'] == cat}
            cat_sums = {k: filtered_sums[k] for k in cat_sheets if k in filtered_sums}

            ranked_sessions = sorted(
                [(filtered_data[k]['country'], s.get('sessions') or 0) for k, s in cat_sums.items()],
                key=lambda x: x[1], reverse=True)
            ranked_plp = sorted(
                [(filtered_data[k]['country'], s.get('plp_conv') or 0) for k, s in cat_sums.items() if s.get('plp_conv') is not None],
                key=lambda x: x[1], reverse=True)
            ranked_duration = sorted(
                [(filtered_data[k]['country'], s.get('duration') or 0) for k, s in cat_sums.items() if s.get('duration')],
                key=lambda x: x[1], reverse=True)

            total_sess = sum(s.get('sessions') or 0 for s in cat_sums.values())
            avg_plp_list = [s['plp_conv'] for s in cat_sums.values() if s.get('plp_conv')]
            avg_plp_cat = sum(avg_plp_list)/len(avg_plp_list) if avg_plp_list else None

            cat_reps[cat] = {
                'total_sessions': total_sess,
                'avg_plp_conv': round(avg_plp_cat, 4) if avg_plp_cat else None,
                'countries_count': len(cat_sheets),
                'ranked_sessions': ranked_sessions[:5],
                'ranked_plp': ranked_plp[:5],
                'ranked_duration': ranked_duration[:5],
            }

        # Executive
        total_sess_all = sum(s.get('sessions') or 0 for s in filtered_sums.values())
        total_sess_prev_all = sum(s.get('sessions_prev') or 0 for s in filtered_sums.values()) if prev else 0
        total_clicks_all = sum(s.get('event_clicks') or 0 for s in filtered_sums.values())
        all_plp = [s['plp_conv'] for s in filtered_sums.values() if s.get('plp_conv') is not None]
        all_purch = [s['purchase_conv'] for s in filtered_sums.values() if s.get('purchase_conv') is not None]
        g_change = ((total_sess_all - total_sess_prev_all) / total_sess_prev_all * 100) if total_sess_prev_all > 0 else None

        top_plp = sorted([(k, s.get('plp_conv', 0)) for k, s in filtered_sums.items() if s.get('plp_conv')], key=lambda x: x[1], reverse=True)
        low_plp = sorted([(k, s.get('plp_conv', 0)) for k, s in filtered_sums.items() if s.get('plp_conv')], key=lambda x: x[1])

        def _clean_key(k):
            # Remove content type prefix for display
            for prefix in ['lineup_', 'feature_library_']:
                if k.startswith(prefix):
                    k = k[len(prefix):]
            return k.replace('_', ' ')

        executive = {
            'latest_month': latest,
            'prev_month': prev,
            'total_sessions': total_sess_all,
            'session_change_pct': round(g_change, 1) if g_change else None,
            'total_clicks': total_clicks_all,
            'avg_plp_conv': round(sum(all_plp)/len(all_plp), 4) if all_plp else None,
            'avg_purchase_conv': round(sum(all_purch)/len(all_purch), 4) if all_purch else None,
            'total_pages_tracked': len(filtered_sums),
            'top_plp_pages': [(_clean_key(k), round(v, 4)) for k, v in top_plp[:3]],
            'low_plp_pages': [(_clean_key(k), round(v, 4)) for k, v in low_plp[:3]],
        }

        return {'executive': executive, 'country_reports': cr, 'category_reports': cat_reps}

    # Build overall + per content type
    overall = _build_reports_for(None)
    per_content_type = {}
    for ct in content_types_found:
        per_content_type[ct] = _build_reports_for(ct)

    return {
        'executive': overall['executive'],
        'country_reports': overall['country_reports'],
        'category_reports': overall['category_reports'],
        'content_types': content_types_found,
        'per_content_type': per_content_type,
        'sheet_summaries': sheet_summaries,
    }


def main():
    print("=" * 60)
    print("LG Buying Guides Dashboard - Data Parser")
    print("=" * 60)

    # Parse master file
    print("\n[1/3] Parsing master link file...")
    pages = parse_master_file()
    print(f"  Found {len(pages)} page entries")

    # Parse monthly data
    print("\n[2/3] Parsing monthly GA data...")
    monthly_data, months = parse_monthly_data()
    print(f"  Parsed {len(monthly_data)} sheets, months: {months}")

    # Compute insights
    print("\n[3/4] Computing insights...")
    insights = compute_insights(monthly_data, months)
    print(f"  Generated {len(insights)} insights")

    # Expert analysis
    print("\n[4/5] Generating expert analysis...")
    expert = compute_expert_analysis(monthly_data, months, pages)
    print(f"  Country reports: {len(expert.get('country_reports', {}))}")
    print(f"  Category reports: {len(expert.get('category_reports', {}))}")

    # Strategic narrative
    print("\n[5/5] Building strategic narrative...")
    narrative = compute_strategic_narrative(monthly_data, months, pages)
    print(f"  Content actions: {len(narrative.get('content_actions', []))}")

    # Build output
    output = {
        'meta': {
            'months_available': months,
            'sheets_parsed': list(monthly_data.keys()),
            'last_updated': None,
            'countries': sorted(set(d['country'] for d in monthly_data.values())),
            'categories': sorted(set(d['category'] for d in monthly_data.values())),
            'content_types': sorted(set(d.get('content_type', 'lineup') for d in monthly_data.values())),
        },
        'pages': pages,
        'monthly_data': {},
        'insights': insights,
        'expert': expert,
        'narrative': narrative,
    }

    # Serialize monthly data (convert for JSON)
    for key, data in monthly_data.items():
        output['monthly_data'][key] = {
            'category': data['category'],
            'country': data['country'],
            'content_type': data.get('content_type', 'lineup'),
            'metrics_2026': data['metrics_2026'],
            'metrics_2025': data['metrics_2025'],
            'insights': data['insights'],
        }

    # Write JSON
    output_path = os.path.join(DATA_DIR, 'data.json')
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2, default=str)

    print(f"\nOutput written to: {output_path}")
    print(f"File size: {os.path.getsize(output_path) / 1024:.1f} KB")
    return output_path


if __name__ == '__main__':
    main()
